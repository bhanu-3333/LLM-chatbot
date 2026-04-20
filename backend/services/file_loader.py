"""
file_loader.py
Universal document loader for multiple file types:
  - PDF  (.pdf)       → PyPDFLoader + OCR fallback
  - Image (.jpg/.jpeg/.png) → pytesseract / easyocr
  - Excel (.xlsx/.xls) → openpyxl / xlrd → table → text
  - Text  (.txt)       → plain read

All loaders return list[Document] compatible with LangChain chunker.
"""

import os
from pathlib import Path
from langchain_core.documents import Document

# Global cache for OCR reader to avoid slow reloading
_EASYOCR_READER = None

# ── Constants ─────────────────────────────────────────────────────────────────
MAX_PAGES      = 100
MIN_CHARS_PAGE = 20

SUPPORTED_EXTENSIONS = {
    ".pdf":  "pdf",
    ".jpg":  "image",
    ".jpeg": "image",
    ".png":  "image",
    ".xlsx": "excel",
    ".xls":  "excel",
    ".txt":  "text",
    ".csv":  "text",
}


# ── Availability checks ───────────────────────────────────────────────────────

def _tesseract_available() -> bool:
    try:
        import pytesseract
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False

def _easyocr_available() -> bool:
    try:
        import easyocr  # noqa
        return True
    except ImportError:
        return False

def _pdf2image_available() -> bool:
    try:
        import pdf2image  # noqa
        return True
    except ImportError:
        return False

def _openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa
        return True
    except ImportError:
        return False

def _xlrd_available() -> bool:
    try:
        import xlrd  # noqa
        return True
    except ImportError:
        return False

def _pil_available() -> bool:
    try:
        from PIL import Image  # noqa
        return True
    except ImportError:
        return False


# ── OCR helpers ───────────────────────────────────────────────────────────────

def _ocr_image_tesseract(img) -> str:
    import pytesseract
    return pytesseract.image_to_string(img)

def _ocr_image_easyocr(img) -> str:
    global _EASYOCR_READER
    import numpy as np
    if _EASYOCR_READER is None:
        import easyocr
        print("[ocr] Initializing EasyOCR Reader (one-time)...")
        _EASYOCR_READER = easyocr.Reader(["en"], gpu=False, verbose=False)
    
    arr = np.array(img)
    lines = _EASYOCR_READER.readtext(arr, detail=0)
    return " ".join(lines)

def _ocr_image(img) -> str:
    """Try Tesseract first, fall back to EasyOCR."""
    if _tesseract_available():
        print("[ocr] Using Tesseract for image")
        return _ocr_image_tesseract(img)
    elif _easyocr_available():
        print("[ocr] Using EasyOCR for image")
        return _ocr_image_easyocr(img)
    else:
        raise ValueError(
            "No OCR engine found. Install pytesseract (+ Tesseract) or easyocr."
        )

def _pdf_to_images(file_path: str, max_pages: int):
    from pdf2image import convert_from_path
    return convert_from_path(file_path, first_page=1, last_page=max_pages, dpi=200)


# ── Individual loaders ────────────────────────────────────────────────────────

def _load_pdf(file_path: str) -> list[Document]:
    """Load PDF with text extraction + OCR fallback for scanned pages."""
    from langchain_community.document_loaders import PyPDFLoader

    file_name = os.path.basename(file_path)
    raw_docs = PyPDFLoader(file_path).load()[:MAX_PAGES]
    text_docs = [d for d in raw_docs if len(d.page_content.strip()) >= MIN_CHARS_PAGE]

    if len(text_docs) == len(raw_docs) and text_docs:
        print(f"[pdf_loader] Text-based PDF — {len(text_docs)} pages")
        return text_docs

    scanned = [i for i, d in enumerate(raw_docs) if len(d.page_content.strip()) < MIN_CHARS_PAGE]
    print(f"[ocr] Detected {len(scanned)} scanned page(s) in PDF")

    if not _pdf2image_available():
        if text_docs:
            print("[ocr] pdf2image not available — returning text-only pages")
            return text_docs
        raise ValueError(
            "PDF contains no extractable text and pdf2image is not installed. "
            "Install poppler + pdf2image for OCR support."
        )

    print(f"[ocr] Converting PDF to images...")
    images = _pdf_to_images(file_path, MAX_PAGES)

    ocr_docs = []
    for page_num, img in enumerate(images):
        text = _ocr_image(img).strip()
        if len(text) < MIN_CHARS_PAGE:
            continue
        ocr_docs.append(Document(
            page_content=text,
            metadata={"source": file_name, "page": page_num, "type": "pdf"}
        ))

    page_map = {d.metadata["page"]: d for d in text_docs}
    for d in ocr_docs:
        page_map[d.metadata["page"]] = d

    final = [page_map[k] for k in sorted(page_map)]
    if not final:
        raise ValueError("Unable to extract text from PDF — both text extraction and OCR failed.")
    return final


def _load_image(file_path: str) -> list[Document]:
    """Extract text from JPEG/PNG using OCR."""
    if not _pil_available():
        raise ValueError("Pillow (PIL) is not installed. Run: pip install Pillow")

    from PIL import Image
    file_name = os.path.basename(file_path)
    img = Image.open(file_path).convert("RGB")
    print(f"[image_loader] Loaded image: {file_name} ({img.size})")

    text = _ocr_image(img).strip()
    if len(text) < MIN_CHARS_PAGE:
        raise ValueError(
            f"No readable text found in image '{file_name}'. "
            "Ensure the image is clear and contains readable text."
        )

    print(f"[image_loader] Extracted {len(text)} characters from image")
    return [Document(
        page_content=text,
        metadata={"source": file_name, "page": 0, "type": "image"}
    )]


def _load_excel(file_path: str) -> list[Document]:
    """Convert Excel sheets to text Documents. Each sheet → one Document."""
    file_name = os.path.basename(file_path)
    ext = Path(file_path).suffix.lower()
    docs = []

    if ext == ".xlsx":
        if not _openpyxl_available():
            raise ValueError("openpyxl is not installed. Run: pip install openpyxl")
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Filter empty rows
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in row_vals):
                    rows.append("\t".join(row_vals))
            text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            if len(text.strip()) >= MIN_CHARS_PAGE:
                docs.append(Document(
                    page_content=text,
                    metadata={"source": file_name, "sheet": sheet_name, "type": "excel"}
                ))

    elif ext == ".xls":
        if not _xlrd_available():
            raise ValueError("xlrd is not installed. Run: pip install xlrd")
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for r in range(ws.nrows):
                row_vals = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                if any(v.strip() for v in row_vals):
                    rows.append("\t".join(row_vals))
            text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            if len(text.strip()) >= MIN_CHARS_PAGE:
                docs.append(Document(
                    page_content=text,
                    metadata={"source": file_name, "sheet": sheet_name, "type": "excel"}
                ))
    else:
        raise ValueError(f"Unsupported Excel extension: {ext}")

    if not docs:
        raise ValueError(f"No readable content found in Excel file '{file_name}'.")

    print(f"[excel_loader] Extracted {len(docs)} sheet(s) from {file_name}")
    return docs


def _load_text(file_path: str) -> list[Document]:
    """Load plain text or CSV files."""
    file_name = os.path.basename(file_path)
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    text = None

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                text = f.read()
            print(f"[text_loader] Read {file_name} with encoding={enc}")
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if text is None:
        raise ValueError(f"Could not decode '{file_name}' with any supported encoding.")

    text = text.strip()
    if len(text) < MIN_CHARS_PAGE:
        raise ValueError(f"Text file '{file_name}' appears to be empty or has insufficient content.")

    print(f"[text_loader] Extracted {len(text)} characters from {file_name}")
    return [Document(
        page_content=text,
        metadata={"source": file_name, "page": 0, "type": "text"}
    )]


# ── Public entry point ────────────────────────────────────────────────────────

def load_file(file_path: str) -> list[Document]:
    """
    Universal file loader. Dispatches to the correct loader based on extension.
    Returns list[Document] ready for chunking.
    """
    ext = Path(file_path).suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(SUPPORTED_EXTENSIONS.keys())
        raise ValueError(
            f"Unsupported file type '{ext}'. Supported types: {supported}"
        )

    kind = SUPPORTED_EXTENSIONS[ext]
    print(f"[file_loader] Loading '{os.path.basename(file_path)}' as type='{kind}'")

    if kind == "pdf":
        return _load_pdf(file_path)
    elif kind == "image":
        return _load_image(file_path)
    elif kind == "excel":
        return _load_excel(file_path)
    elif kind == "text":
        return _load_text(file_path)
    else:
        raise ValueError(f"Unknown file kind: {kind}")
